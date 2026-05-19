import streamlit as st
import pandas as pd
import re
import io
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# === 【V9.16 终极抗干扰疫苗】===
# 强行修复 openpyxl 最新版读取带下拉菜单的 Excel 时的底层 'id' 报错 BUG
try:
    import openpyxl.worksheet.datavalidation
    _orig_init = openpyxl.worksheet.datavalidation.DataValidation.__init__
    def _safe_init(self, *args, **kwargs):
        kwargs.pop('id', None) # 拔掉引发报错的毒牙
        _orig_init(self, *args, **kwargs)
    openpyxl.worksheet.datavalidation.DataValidation.__init__ = _safe_init
except Exception:
    pass
# ===============================

# 1. 网页全局配置
st.set_page_config(page_title="饱里宝气 | 极简分单系统 v9.16", page_icon="🍱", layout="wide")

st.title("🍱 饱气极简分单中枢 V9.16 云端版")
st.markdown("V9.16 更新：**bj版图扩容与云端重构** | 注入表单解析疫苗，彻底解决云端环境冲突。")

# 2. 核心分单逻辑 (绝对优先级排序引擎)
def categorize_final(addr):
    if pd.isna(addr): return '需人工核对'
    addr = str(addr).replace('\n', '')
    
    if any(k in addr for k in ['中医药', '比亚迪', '奥克斯', '交投', '含浦', '湖大子弟小学', '湖南大学子弟小学', '798', '科学村', '国际艺术园区', '四张机']): return '达达配送片区'
    if any(k in addr for k in ['金域缇香', '靳江', '黄鹤', '惟盛园', '中南壹号', '中南一号', '湘桥', '龙湖冠寓', '麓山公寓', '阳光100']): return '毛哥'
    
    if any(k in addr for k in ['德智', '后湖小区', '后湖门', '11栋围栏', '生命交叉', '生命医学交叉', '天马小区', '天马南门', '天马学生公寓(南门)', '天马学生公寓（南门）']): return 'ds'
    
    if any(k in addr for k in ['天马公寓', '天马学生公寓', '天马大酒店', '生物学院', '笔芯', '未来乡村', '国家电能', '天马基地', '新东方']): return 'bj'
    
    if any(k in addr for k in ['桃花坪', '桃花公寓', '南院', '美术楼', '和乐楼', '体育馆', '矿冶院']): return 'fx'
    
    if any(k in addr for k in ['湖南师范大学', '师大', '二里半', '研究生2舍', '研究生3舍', '研究生5舍', '江边', '世承', '木兰楼', '科教院', '九栋', '十栋', '附中', '木兰', '学堂坡', '景德', '法学院', '麓枫和苑']):
        if any(k in addr for k in ['师大2', '格物', '理学院', '青商众创']): return '骑手' 
        return 'zw'

    if any(k in addr for k in ['湖南大学', '湖大', '工商管理', '图书馆总图', '湖大总图', '十七舍', '中楼', '十三舍', '十八宿', '十四舍', '空天科学']): return 'ab'
    if any(k in addr for k in ['本部', '科技楼', '民主楼', '管理楼', '采矿楼', '第一办公楼', '地学楼', '化学楼', '生物楼', '粉末冶金', '米塔尔', '科教南楼', '图书馆', '2舍', '3舍', '5舍', '15舍', '三一大楼']): return '本部'
    if any(k in addr for k in ['升华', '南6', '南六', '南八', '南五', '青年教师', '教职工17', '逸间', '麓南校区']): return '升华'
    if any(k in addr for k in ['新校区', '江湾楼', '德必', '半导体', 'MeUMe', '腾讯', '潇湘', '慧博云通', 'A座', 'B座', '外国语', '校史馆', '信息楼', '化学', '数学', '物理']): return '新校区'

    return '需人工核对'

# 3. 路线优先级映射
ADDRESS_KEYWORD_ORDER = {
    '新校区': ['1中南', '2中南', '3中南', '4中南', '6中南', '8中南', '化学', '数学', '物理', '信息', '机电', '江湾', '德必', '半导体', 'MeUMe', '腾讯', '潇湘', '慧博云通'],
    'ds': ['1湖南', '2湖南', '3湖南', '4湖南', '5天马', '6后湖'], 
    '本部': [f"{i}中南" for i in range(1, 18)],
    'ab': [f"{i}湖南" for i in range(1, 20)], 
    '升华': [f"{i}中南" for i in range(1, 36)],
    'fx': ['1中南', '2湖南', '3湖南', '美术', '和乐', '矿冶'],
    'zw': [f"{i}师大" for i in range(1, 25)] + [f"{i}湖南" for i in range(1, 25)] + ['世承', '木兰', '老化工', '学堂坡', '江边', '景德', '麓枫和苑'],
    '骑手': ['格物楼', '理学院', '生命科学', '青商众创'],
    'bj': ['天马公寓', '天马学生', '天马大酒店', '生物学院', '未来乡村', '笔芯', '国家电能', '天马基地', '新东方'] 
}

# 4. 地址标准化引擎
def apply_prefix(orig_addr, addr, person):
    clean_addr = str(addr).strip()
    clean_addr = clean_addr.replace('湖南师范大学国际艺术园区', '国际艺术园区')
    clean_addr = clean_addr.replace('中南大学新校区化学天马小区', '天马小区').replace('化学天马小区', '天马小区')
    orig_str = clean_addr.strip()
    
    clean_addr = re.sub(r'^\d+(?=(湖南|中南|师大|万科|天马|后湖|国家|麓枫|世承|木兰|研究生|江边|桃花|美术|和乐|矿冶|德智|生命|国际艺术园区|新东方))', '', clean_addr)
    clean_addr = re.sub(r'(麻烦送到|麻烦送|送到|麻烦|请送|请送到)', '', clean_addr)
    
    company_community_keywords = ['公司', '研究院', '小区', '德必', 'MeUMe', '慧博云通', '腾讯', '半导体', '麓枫和苑', '天马基地', '未来乡村', '惟盛园', '中南壹号', '外卖架', '婚礼', '国际艺术园区', '绸缎庄', '新东方']
    if any(k in orig_str for k in company_community_keywords):
        clean_addr = re.sub(r'(中南大学新校区|中南大学校本部|中南大学升华公寓|中南大学|湖南师范大学|湖南师大|湖师大|湖南大学|湖大|新校区|校本部|本部)', '', clean_addr)
    else:
        clean_addr = clean_addr.replace('中南大学校本部', '').replace('中南大学新校区', '').replace('中南大学升华公寓', '').replace('中南大学', '').replace('本部', '')
    
    if person == '达达配送片区':
        if '国际艺术园区' in orig_str or '四张机' in orig_str:
            clean_addr = clean_addr.replace('湖南师范大学', '').replace('师大', '')
            clean_addr = clean_addr.strip() 
        return clean_addr

    if person == '毛哥':
        if any(k in orig_str or k in clean_addr for k in ['中南壹号', '中南一号']):
            if not any(k in clean_addr for k in ['中南壹号', '中南一号']): clean_addr = "中南壹号院" + clean_addr
        elif '惟盛园' in orig_str or '惟盛园' in clean_addr:
            if '惟盛园' not in clean_addr: clean_addr = "惟盛园" + clean_addr
        elif '金域缇香' in orig_str or '金域缇香' in clean_addr:
            if '金域缇香' not in clean_addr: clean_addr = "万科·金域缇香" + clean_addr
        return clean_addr

    if person == 'bj':
        if '国家电能' in orig_str or '天马基地' in orig_str or '国家电能' in clean_addr or '天马基地' in clean_addr: 
            return "国家电能变换与控制工程技术研究中心(天马基地)"
        return clean_addr

    if person == 'fx':
        clean_addr = clean_addr.replace('湖南师范大学', '').replace('湖南师大', '').replace('湖师大', '').replace('师大', '')
        if '体育馆' in orig_str or '体育馆' in clean_addr: return "1中南大学体育馆地铁口铁门处"
        if '桃花坪图书馆' in orig_str or '桃花坪图书馆' in clean_addr: return "3湖南师范大学桃花坪图书馆"
        return f"湖南师范大学{clean_addr}"

    if person in ['zw', '骑手']:
        clean_addr = clean_addr.replace('湖南师范大学', '').replace('湖南师大', '').replace('湖师大', '').replace('师大', '').replace('二里半校区', '').replace('二里半', '')
        clean_addr = re.sub(r'^\d+', '', clean_addr).strip()
        
        if "九栋" in orig_str or "九栋" in clean_addr: return "12师大二里半校区九栋"
        if "十栋" in orig_str or "十栋" in clean_addr: return "10师大二里半校区十栋"
        z_map = {"木兰楼": 4, "师大附中停车场": 15, "附中停车场": 15, "科教院": 18, "世承": 1, "木兰": 4, "研究生2舍": 13, "研究生3舍": 14, "江边3舍": 14, "麓枫和苑": 0}
        for k, v in z_map.items():
            if k in orig_str or k in clean_addr:
                if k == "师大附中停车场" or k == "附中停车场": return f"15湖南师范大学师大附中停车场门口"
                if k == "木兰楼": return f"4湖南师范大学二里半校区木兰楼"
                if k == "科教院": return f"18湖南师范大学二里半校区科教院"
                if k == "麓枫和苑":
                    if "麓枫和苑" not in clean_addr: return "麓枫和苑" + clean_addr
                    return clean_addr[clean_addr.find("麓枫和苑"):]
                if k not in clean_addr: clean_addr = k + clean_addr
                return f"{v}湖南师范大学{clean_addr}"
        if '湖南师范大学' not in clean_addr: return f"湖南师范大学{clean_addr}"
        return clean_addr

    if person == 'ab':
        clean_addr = clean_addr.replace('湖南大学', '').replace('湖大', '')
        if '图书馆总图' in orig_str or '湖大总图' in orig_str or '图书馆总图' in clean_addr or '湖大总图' in clean_addr: return "2湖南大学图书馆总图"
        if '工商管理' in orig_str or '工商管理' in clean_addr: return "1湖南大学工商管理学院"
        return f"湖南大学{clean_addr}"

    if person == 'ds':
        clean_addr = clean_addr.replace('湖南大学', '').replace('中南大学', '').replace('湖大', '').replace('化学化工学院', '').replace('化学', '')
        if '11栋围栏' in orig_str or '11栋围栏' in clean_addr: return "1湖南大学德智园11栋围栏"
        if any(k in orig_str or k in clean_addr for k in ['后湖门', '德智园学生公寓', '德智后湖']): return "2湖南大学德智园学生公寓(后湖门)"
        if any(k in orig_str or k in clean_addr for k in ['天马南门', '学生公寓(南门)', '学生公寓（南门）']): return "3湖南大学天马学生公寓(南门)"
        if any(k in orig_str or k in clean_addr for k in ['生命交叉', '生命医学交叉']): return "4湖南大学生命医学交叉研究院"
        
        if '天马小区' in orig_str or '天马' in clean_addr:
            if '天马小区' not in clean_addr: clean_addr = "天马小区" + clean_addr.replace('天马', '')
            clean_addr = clean_addr[clean_addr.find("天马小区"):]
            return f"5{clean_addr}"
            
        if '后湖小区' in orig_str or '后湖' in clean_addr:
            if '后湖小区' not in clean_addr: clean_addr = "后湖小区" + clean_addr.replace('后湖', '')
            clean_addr = clean_addr[clean_addr.find("后湖小区"):]
            return f"6{clean_addr}"
        return clean_addr

    if person == '本部':
        clean_addr = re.sub(r'科教南楼.*?实际地址为三一大楼', '三一大楼', clean_addr)
        clean_addr = clean_addr.replace('实际地址为三一大楼', '三一大楼')
        b_map = {"科技楼":1, "图书馆":2, "民主楼":3, "管理楼":4, "采矿楼":5, "第一办公楼":6, "地学楼":7, "化学楼":8, "生物楼":9, "粉末冶金":10, "米塔尔":11, "科教南楼":12, "2舍":13, "3舍":14, "5舍":15, "15舍":16, "三一大楼":17}
        for k, v in b_map.items():
            if k in orig_str or k in clean_addr:
                if k not in clean_addr: clean_addr = k + clean_addr
                return f"{v}中南大学校本部{clean_addr}"
        return f"中南大学校本部{clean_addr}"

    if person == '升华':
        clean_addr = clean_addr.replace('麓南校区', '').replace('升华学生公寓', '')
        clean_addr = re.sub(r'(麓南校区)?南6舍?', '南六舍', clean_addr)
        clean_addr = clean_addr.replace('南6舍', '南六舍').replace('南6', '南六舍').replace('南六', '南六舍').replace('舍舍', '舍')
        
        if '青年教师公寓2栋' in orig_str or '青年教师公寓2栋' in clean_addr:
            suffix = clean_addr.split('青年教师公寓2栋')[-1] if '青年教师公寓2栋' in clean_addr else ''
            return f"27南校区青年教师公寓2栋{suffix}"
        if '青年教师公寓3栋' in orig_str or '青年教师公寓3栋' in clean_addr:
            suffix = clean_addr.split('青年教师公寓3栋')[-1] if '青年教师公寓3栋' in clean_addr else ''
            return f"28南校区青年教师公寓3栋{suffix}"

        s_map = {"14栋":1, "19栋":2, "18栋":3, "13栋":4, "7栋":5, "6栋":6, "5栋":7, "1栋":8, "2栋":9, "10栋":10, "11栋":11, "17栋":12, "南八":13, "南五":14, "南六舍":15, "20栋":16, "21栋":17, "24栋":18, "25栋":19, "26栋":20, "29栋":21, "15栋":22, "16栋":23, "27栋":24, "28栋":25, "31栋":26, "青年教师公寓2栋":27, "青年教师公寓3栋":28, "32栋":29, "33栋":30, "34栋":31, "38":32, "35栋":33, "教职工17栋":34, "逸间":35}
        for k, v in s_map.items():
            if k in orig_str or k in clean_addr:
                if k not in clean_addr: clean_addr = k + clean_addr
                return f"{v}中南大学升华公寓{clean_addr}"
        return f"中南大学升华公寓{clean_addr}"

    if person == '新校区':
        pure_companies = ['慧博云通', '德必', '半导体', 'MeUMe', '腾讯', '潇湘', '公司', '研究院', '外卖架', '婚礼']
        if any(c in orig_str for c in pure_companies): return clean_addr
        n_map = {"A座":1, "外国语":2, "B座":3, "图书馆":4, "化学":6, "数学":8, "物理":0, "信息":0}
        for k, v in n_map.items():
            if k in orig_str or k in clean_addr:
                if k == "A座": return "1中南大学新校区A座_不让放放在B座"
                if k not in clean_addr: clean_addr = k + clean_addr
                return f"{v}中南大学新校区{clean_addr}" if v != 0 else f"中南大学新校区{clean_addr}"
        return f"中南大学新校区{clean_addr}"

    return clean_addr

# 5. 动线排序权重引擎
def get_address_rank(person, addr):
    addr_str = str(addr)
    if person == 'ds' and '天马小区' in addr_str:
        match = re.search(r'天马小区(\d+)', addr_str)
        if match: return 5.0 + (1000 - int(match.group(1))) / 10000
    if person in ADDRESS_KEYWORD_ORDER:
        for idx, kw in enumerate(ADDRESS_KEYWORD_ORDER[person]):
            if addr_str.startswith(kw): return float(idx)
            elif kw in addr_str and "中南" not in kw and "湖南" not in kw and "师大" not in kw: return float(idx)
    return 999.0

# 6. 常规功能引擎与【天眼级备注抓取】
def optimize_remark(remark):
    if pd.isna(remark) or str(remark).strip() == '': return '无'
    opt = str(remark).replace('beef', '牛肉').replace('Beef', '牛肉').replace('shrimp', '虾滑').replace('Shrimp', '虾滑')
    keywords = ['不要洋葱', '去洋葱', '换牛肉', '换虾滑', '不要辣', '免辣', '加量', '多饭', '少饭', '多菜', '体验餐']
    for kw in keywords:
        if kw in opt: opt = opt.replace(kw, f"【{kw}】")
    return opt

def clean_address(addr):
    if pd.isna(addr): return ''
    cleaned = re.sub(r'(湖南省|长沙市|岳麓区)\s*', '', str(addr))
    return cleaned.strip()

def smart_refine_address(original_addr, remark):
    remark_str = str(remark).replace('\n', '') if pd.notna(remark) else ''
    if any(k in remark_str for k in ['舍', '栋', '楼', '公寓', '宿舍', '院', '馆', '中心', '园', '小区', '村', '座']):
        clean_rem = re.sub(r'(麻烦送到|麻烦送|麻烦|请送|请送到|放附近|改送到|送到|改送|送去|实际地址为|实际地址是|地址改为|地址为|地址是|谢谢|啦)', '', remark_str)
        pattern = r'([^\s,，。!！\-\+]{2,30}(?:舍|栋|楼|公寓|宿舍|研究生\d+舍|院|馆|中心|园|小区|村|门|座|研究院|学院|公司|外卖架|婚礼|园区|绸缎庄|新东方)[^\s,，。!！\-\+]*)'
        match = re.search(pattern, clean_rem)
        if match: return match.group(1)
    return clean_address(original_addr)

def smart_tagger(remark, current_tag):
    r = str(remark).strip() if pd.notna(remark) else ''
    t = str(current_tag).strip() if pd.notna(current_tag) else ''
    combined = r + " " + t
    if '体验餐' in combined: return '体验餐'
    if any(kw in combined for kw in ['加量', '多饭']): return "加量"
    if any(kw in combined for kw in ['少饭多菜', '多蔬菜', '少饭']): return "少饭多菜"
    if any(kw in combined for kw in ['不辣', '无辣', '不要辣', '免辣']): return "免辣"
    return t if t else ''

def apply_remark_overrides(addr, rem):
    if pd.isna(rem): return addr
    rem_str = str(rem)
    if '南六舍' in rem_str: return '南六舍'
    if '科技楼' in rem_str: return '科技楼'
    if '三一大楼' in rem_str: return '三一大楼'
    return addr

def clean_remark_overrides(rem):
    if pd.isna(rem): return rem
    rem_str = str(rem)
    if '南六舍' in rem_str:
        res = rem_str.replace('放附近南六舍谢谢啦', '').replace('送到南六舍', '').replace('南六舍', '').strip()
        return res if res else '无'
    if '科技楼' in rem_str:
        res = rem_str.replace('科技楼', '').strip()
        return res if res else '无'
    if '三一大楼' in rem_str:
        res = rem_str.replace('实际地址为三一大楼', '').replace('三一大楼', '').strip()
        return res if res else '无'
    return rem_str

# 7. 文件处理主程序
uploaded_file = st.file_uploader("📂 请上传原始订单表 (验证无敌云端版)", type=['csv', 'xlsx', 'xls'])

if uploaded_file is not None:
    try:
        df = pd.read_csv(uploaded_file) if uploaded_file.name.endswith('.csv') else pd.read_excel(uploaded_file)
        mapping = {'用户昵称': next((c for c in df.columns if any(x in c for x in ['昵称', '姓名', '客户'])), '用户昵称'),
                   '数量': next((c for c in df.columns if any(x in c for x in ['实际配送', '数量', '份数', '单量'])), '数量'),
                   '配送地址': next((c for c in df.columns if '地址' in c), '配送地址'),
                   '订单备注': next((c for c in df.columns if '备注' in c), '订单备注'),
                   '订单标签': next((c for c in df.columns if '标签' in c), '订单标签')}

        with st.spinner("🚀 正在极速处理订单，云端防卡死引擎已启动..."):
            
            df = df.dropna(subset=[mapping['用户昵称'], mapping['配送地址']], how='all')
            df[mapping['配送地址']] = df.apply(lambda x: apply_remark_overrides(x[mapping['配送地址']], x[mapping['订单备注']]), axis=1)
            df[mapping['订单备注']] = df[mapping['订单备注']].apply(clean_remark_overrides)

            df['__refined_addr__'] = df.apply(lambda x: smart_refine_address(x[mapping['配送地址']], x[mapping['订单备注']]), axis=1)
            df['配送员'] = df['__refined_addr__'].apply(categorize_final)
            df[mapping['配送地址']] = df.apply(lambda x: apply_prefix(x['__refined_addr__'], x['__refined_addr__'], x['配送员']), axis=1)
            
            df[mapping['订单标签']] = df.apply(lambda x: smart_tagger(x[mapping['订单备注']], x[mapping['订单标签']]), axis=1)
            df[mapping['订单备注']] = df[mapping['订单备注']].apply(optimize_remark)
            
            order_list = ['新校区', 'ds', '本部', '毛哥', 'fx', 'zw', '骑手', 'ab', 'bj', '升华', '达达配送片区']
            df['PersonRank'] = df['配送员'].apply(lambda x: order_list.index(x) if x in order_list else 99)
            df['AddressRank'] = df.apply(lambda x: get_address_rank(x['配送员'], x[mapping['配送地址']]), axis=1)
            
            df_sorted = df.sort_values(by=['PersonRank', 'AddressRank', mapping['配送地址']], ascending=[True, True, True])

            final_cols = [mapping['用户昵称'], mapping['数量'], mapping['订单标签'], mapping['配送地址'], '配送员', mapping['订单备注']]
            df_final = df_sorted[final_cols]
            df_final.columns = ['用户昵称', '数量', '订单标签', '配送地址', '配送员', '订单备注']

            st.success("✅ V9.16 处理完成！饱气云端中枢已彻底免疫表格结构 BUG！")
            st.dataframe(df_final.head(10), use_container_width=True)
            
            output = io.BytesIO()
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                df_final.to_excel(writer, index=False, sheet_name='饱气打印单')
                ws = writer.sheets['饱气打印单']
                
                h_fill = PatternFill(start_color="4CAF50", end_color="4CAF50", fill_type="solid")
                h_font = Font(bold=True, size=14, color="FFFFFF")
                
                for cell in ws[1]:
                    cell.fill, cell.font = h_fill, h_font
                    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                for row in ws.iter_rows(min_row=2):
                    for cell in row:
                        cell.font = Font(size=11)
                        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                
                for col in ws.columns:
                    max_length = 0
                    c_letter = col[0].column_letter
                    for cell in col:
                        try:
                            if cell.value is not None:
                                length = sum(2.2 if ord(c) > 127 else 1.2 for c in str(cell.value))
                                if length > max_length: max_length = length
                        except: pass
                    ws.column_dimensions[c_letter].width = max_length + 4
                    
                ws.row_dimensions[1].height = 25
            
            st.download_button(label="⬇️ 下载饱气精排打印版", data=output.getvalue(),
                               file_name=f"饱气精排版_V9.16.xlsx",
                               mime='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    except Exception as e:
        st.error(f"处理失败: {e}")
